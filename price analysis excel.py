# ExcelファイルのデータをPythonで読み込んで加工、新しいExcelファイルに保存
from openpyxl import load_workbook # openpyxlはExcelを読み書きするためのライブラリ
# ドキュメント https://openpyxl.readthedocs.io/en/stable/
from collections import defaultdict
from typing import Dict, Tuple

# 型の定義
BrandGenderKey = Tuple[str, str]
BrandGenderData = Dict[BrandGenderKey, Dict[str, float]]

# Excelファイルを読み込む
# openpyxlで定義されているクラス Workbookクラス:ワークブック全体、Worksheetクラス:一つのシート
file_path = "shoes_data.xlsx"
wb = load_workbook(file_path) # ファイル全体を読んでいる
#print(wb.sheetnames) # シート名一覧のリストが表示される
ws = wb.active # ワークブックにあるアクティブ中のワークシート(現在操作可能なシート)を取得
# ws1=wb['Sheet1'],ws2=wb['Sheet2']とシートを表す変数を指定し、wb.active=ws2でアクティブ化
# for ws in wb.worksheets: すべてのシートの選択状態を解除(これをやってから1つだけ選択)
#    ws.sheet_view.tabSelected = False

# メモリ上にbrand_gender_dataを作成。変数:型の定義 で型の宣言を行いながら変数を定義
brand_gender_data: BrandGenderData = \
    defaultdict(lambda: {"total_price": 0.0, "total_size": 0.0, "count": 0})
# defaultdict(lambda: ):存在しないキーが参照されたときに自動的にデフォルト値を生成
# if key not in brand_gender_data:brand_gender_data[key]={デフォルト設定}と書かなくていい
# lambdaは匿名関数(無名関数)を作成するための構文

# Excelからデータを読み込みbrand_gender_dataに収納
for row in ws.iter_rows(min_row=2, values_only=True): # iter_rows():Excelからデータを取得
# min_row=2はExcelの2(値の上から2番目)、max_low,min_col,max_col
# values_only=True:セルの値のみを取得、=False:値だけでなくセル自体(スタイルを含む)を取得
    brand, gender, price, size = row
    key = (brand, gender)
    brand_gender_data[key]["total_price"] += price
    brand_gender_data[key]["total_size"] += size
    brand_gender_data[key]["count"] += 1

#print(brand_gender_data) # 理解のため表示

# 新しいシート"Grouped Data"(result_ws)を作成。
result_ws = wb.create_sheet("Grouped Data")
result_ws.append(["Brand", "Gender", "Average Price", "Average Size", "Model Count"])

# ブランドをソート
brands = sorted(set(brand for brand, _ in brand_gender_data.keys()))

# 新しいシートresult_wsにbrand_gender_dataに収納したデータを書き込む
for brand in brands:
    for gender in ["M", "W"]:  # 性別順序を指定
        key = (brand, gender)
        if key in brand_gender_data:
            data = brand_gender_data[key]
            count = data["count"]
            avg_price = data["total_price"] / count
            avg_size = data["total_size"] / count
            result_ws.append([brand, gender, avg_price, avg_size, count])

# 保存前に処理結果を表示させて確認
for row in result_ws.iter_rows(values_only=True):
    print(row)

# 保存
wb.save("shoe_price_analysis.xlsx")